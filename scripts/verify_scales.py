# -*- coding: utf-8 -*-
import sys, io, os, zipfile
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

MXL_DIR = 'prisma/data/mxl'
EXCEL = 'prisma/data/arcoda_scales_936.xlsx'

wb = openpyxl.load_workbook(EXCEL)
ws = wb[wb.sheetnames[0]]

BOW_KEY_MAP = {"デタシェ": "detache", "スタッカート": "staccato", "スラー2音": "slur2",
               "スラー4音": "slur4", "スラー8音": "slur8", "レガート": "legato"}
SCALE_TYPE_MAP = {"長音階": "major", "自然的短音階": "natural_minor", "和声的短音階": "harmonic_minor",
                  "旋律的短音階": "melodic_minor", "半音階": "chromatic", "短音階": ""}

issues = []
note_count_by_type = {}

for row_idx in range(2, ws.max_row + 1):
    title = ws.cell(row_idx, 1).value
    if not title:
        continue
    root = ws.cell(row_idx, 4).value or "C"
    mode = ws.cell(row_idx, 5).value or "major"
    scale_type_ja = ws.cell(row_idx, 6).value or "長音階"
    minor_variant = ws.cell(row_idx, 7).value or ""
    octaves = int(ws.cell(row_idx, 8).value or 1)
    bow = ws.cell(row_idx, 9).value or "デタシェ"
    bow_key = BOW_KEY_MAP.get(bow, "detache")
    # variant_key: use minor_variant (col G) for minor, "chromatic" for chromatic
    variant_key = ""
    if mode == "minor":
        variant_key = (minor_variant or "").strip().lower()
    if scale_type_ja == "半音階":
        variant_key = "chromatic"

    filename = f"scale_{row_idx:04d}_{root}_{mode}_{variant_key}_{octaves}oct_{bow_key}.mxl"
    filepath = os.path.join(MXL_DIR, filename)

    if not os.path.exists(filepath):
        issues.append(f"MISSING: {filename}")
        continue

    with zipfile.ZipFile(filepath) as zf:
        xml = zf.read('score.xml').decode('utf-8')
    note_count = xml.count('<pitch>')

    type_key = f"{scale_type_ja}|{minor_variant}|{octaves}oct"
    if type_key not in note_count_by_type:
        note_count_by_type[type_key] = []
    note_count_by_type[type_key].append(note_count)

    if note_count < 8:
        issues.append(f"SHORT ({note_count} notes): {filename} - {title}")

print("=== Note counts by scale type ===")
for key in sorted(note_count_by_type.keys()):
    counts = note_count_by_type[key]
    min_c = min(counts)
    max_c = max(counts)
    print(f"  {key}: min={min_c}, max={max_c}, files={len(counts)}")

print(f"\n=== Issues: {len(issues)} ===")
for issue in issues[:30]:
    print(f"  {issue}")
