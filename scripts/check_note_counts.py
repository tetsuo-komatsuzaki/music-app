# -*- coding: utf-8 -*-
import sys, io, zipfile, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import os

MXL_DIR = 'prisma/data/mxl'

short_files = []
for filename in sorted(os.listdir(MXL_DIR)):
    if not filename.endswith('.mxl'):
        continue
    filepath = os.path.join(MXL_DIR, filename)
    with zipfile.ZipFile(filepath) as zf:
        xml = zf.read('score.xml').decode('utf-8')
    # Count <note> elements that are NOT rests
    note_count = xml.count('<pitch>')
    if note_count <= 6:
        short_files.append((filename, note_count))

print(f"Files with <= 6 notes: {len(short_files)}")
for f, c in short_files[:20]:
    print(f"  {c} notes: {f}")

# Also show note counts for specific scale types
import openpyxl
wb = openpyxl.load_workbook('prisma/data/arcoda_scales_936.xlsx')
ws = wb[wb.sheetnames[0]]
scale_types = set()
for r in range(2, ws.max_row + 1):
    st = ws.cell(r, 6).value
    variant = ws.cell(r, 7).value
    oct = ws.cell(r, 8).value
    scale_types.add((st, variant, oct))

print(f"\nUnique scale type/variant/octave combos:")
for st, v, o in sorted(scale_types):
    print(f"  {st} | {v} | {o}oct")
