# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

wb = openpyxl.load_workbook('prisma/data/arcoda_scales_936.xlsx')
print('Sheets:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
print('Rows:', ws.max_row, 'Cols:', ws.max_column)

headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
print('Headers:', headers)

for r in range(2, 5):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    print(f'Row {r}: {row}')
